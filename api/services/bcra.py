"""BCRA API helpers and caching logic extracted from the main Flask entrypoint."""

from __future__ import annotations

import io
import re
import time
import unicodedata
import warnings
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from typing import (
    Any,
    Callable,
    Dict,
    Iterable,
    List,
    Mapping,
    Optional,
    Tuple,
    TypeVar,
    Union,
    cast,
)

import redis
import requests
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning

from api.services.redis_helpers import redis_get_json, redis_set_json, redis_setex_json
from api.utils import (
    fmt_num,
    local_cache_get,
    now_utc_iso,
    parse_date_string,
    parse_monetary_number,
    to_ddmmyy,
    to_es_number,
    update_local_cache,
)


CachedRequestFn = Callable[..., Optional[Dict[str, Any]]]
RedisFactoryFn = Callable[..., redis.Redis]
AdminReporterFn = Callable[[str, Optional[Exception], Optional[Dict[str, Any]]], None]
CacheHistoryFn = Callable[[int, str, redis.Redis], Optional[Dict[str, Any]]]


TTL_BCRA = 300  # 5 minutes
TTL_COUNTRY_RISK = 300  # 5 minutes for BondTerminal endpoint
TTL_MAYORISTA_MISSING = 300  # 5 minutes sentinel for missing mayorista values
CACHE_STALE_GRACE_BCRA = 6 * 3600  # allow showing last BCRA data up to 6h stale
CACHE_STALE_GRACE_BANDS = 3600  # currency bands fall back for 1h


_cached_requests_fn: Optional[CachedRequestFn] = None
_redis_factory_fn: Optional[RedisFactoryFn] = None
_admin_reporter_fn: Optional[AdminReporterFn] = None
_cache_history_fn: Optional[CacheHistoryFn] = None


BA_TZ = timezone(timedelta(hours=-3))


def _normalize_text(value: Any) -> str:
    """Return lowercase ASCII-normalized text for fuzzy comparisons."""

    try:
        text = str(value or "")
    except Exception:
        text = ""
    normalized = (
        unicodedata.normalize("NFKD", text)
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    return normalized.lower()


_bcra_local_cache: Dict[str, Any] = {
    "value": None,
    "expires_at": 0.0,
    "stale_until": 0.0,
    "meta": {},
}
_bcra_failure_until: float = 0.0

_currency_band_local_cache: Dict[str, Any] = {
    "value": None,
    "expires_at": 0.0,
    "stale_until": 0.0,
    "meta": {},
}
_currency_band_failure_until: float = 0.0


def _get_bcra_failure_until() -> float:
    return _bcra_failure_until


def _set_bcra_failure_until(value: float) -> None:
    global _bcra_failure_until
    _bcra_failure_until = value


def _get_currency_band_failure_until() -> float:
    return _currency_band_failure_until


def _set_currency_band_failure_until(value: float) -> None:
    global _currency_band_failure_until
    _currency_band_failure_until = value


def configure(
    *,
    cached_requests: CachedRequestFn,
    redis_factory: RedisFactoryFn,
    admin_reporter: Optional[AdminReporterFn] = None,
    cache_history: Optional[CacheHistoryFn] = None,
) -> None:
    """Register external dependencies required by the BCRA helpers."""

    global _cached_requests_fn, _redis_factory_fn, _admin_reporter_fn, _cache_history_fn

    _cached_requests_fn = cached_requests
    _redis_factory_fn = redis_factory
    _admin_reporter_fn = admin_reporter
    _cache_history_fn = cache_history


def reset_local_caches() -> None:
    """Reset in-memory caches for testing."""

    global _bcra_failure_until, _currency_band_failure_until

    _bcra_local_cache.update(
        {"value": None, "expires_at": 0.0, "stale_until": 0.0, "meta": {}}
    )
    _bcra_failure_until = 0.0

    _currency_band_local_cache.update(
        {"value": None, "expires_at": 0.0, "stale_until": 0.0, "meta": {}}
    )
    _currency_band_failure_until = 0.0


def _require_configured(func_name: str) -> None:
    if _cached_requests_fn is None or _redis_factory_fn is None:
        raise RuntimeError(f"BCRA service not configured before calling {func_name}")


def _call_cached_requests(*args, **kwargs):
    _require_configured("cached_requests")
    assert _cached_requests_fn is not None
    return _cached_requests_fn(*args, **kwargs)


def _config_redis(*args, **kwargs) -> redis.Redis:
    _require_configured("config_redis")
    assert _redis_factory_fn is not None
    return _redis_factory_fn(*args, **kwargs)


def _admin_report(message: str, error: Optional[Exception] = None, extra: Optional[Dict[str, Any]] = None) -> None:
    if _admin_reporter_fn:
        _admin_reporter_fn(message, error, extra)


def _get_cache_history(hours_ago: int, cache_key: str, client: redis.Redis) -> Optional[Dict[str, Any]]:
    if hours_ago <= 0:
        return None
    if _cache_history_fn is None:
        return None
    return _cache_history_fn(hours_ago, cache_key, client)


def _load_cached_entry(
    *,
    cache_key: str,
    local_cache: Dict[str, Any],
    ttl: int,
    stale_grace: int,
    allow_stale: bool,
    redis_error_message: Optional[str] = None,
    require_label: Optional[str] = None,
) -> Tuple[Optional[Dict[str, Any]], Dict[str, Any]]:
    _require_configured(require_label or cache_key)

    redis_client: Optional[redis.Redis]
    try:
        redis_client = _config_redis()
    except Exception as exc:
        if redis_error_message:
            print(f"{redis_error_message}: {exc}")
        redis_client = None

    if redis_client is not None:
        cached = redis_get_json(redis_client, cache_key)
        if cached:
            payload: Dict[str, Any]
            if isinstance(cached, dict) and "data" in cached:
                payload = cast(Dict[str, Any], cached)
            else:
                payload = {"data": cached}
            fetched_at = cast(Optional[str], payload.get("fetched_at"))
            update_local_cache(
                local_cache,
                cast(Dict[str, Any], payload["data"]),
                ttl,
                stale_grace,
                fetched_at,
            )
            return cast(Dict[str, Any], payload["data"]), {
                "is_fresh": True,
                "fetched_at": fetched_at,
            }

        if allow_stale:
            last_success = redis_get_json(redis_client, f"{cache_key}:last_success")
            if isinstance(last_success, dict) and "data" in last_success:
                fetched_at = cast(Optional[str], last_success.get("fetched_at"))
                update_local_cache(
                    local_cache,
                    cast(Dict[str, Any], last_success["data"]),
                    0,
                    stale_grace,
                    fetched_at,
                )
                return cast(Dict[str, Any], last_success["data"]), {
                    "is_fresh": False,
                    "fetched_at": fetched_at,
                }

    local_value, is_fresh, meta = local_cache_get(local_cache, allow_stale=allow_stale)
    if local_value is not None:
        return cast(Dict[str, Any], local_value), {
            "is_fresh": is_fresh,
            "fetched_at": meta.get("fetched_at"),
        }

    return None, {"is_fresh": False, "fetched_at": None}


def _persist_cache_entry(
    *,
    cache_key: str,
    data: Dict[str, Any],
    local_cache: Dict[str, Any],
    ttl: int,
    stale_grace: int,
    on_redis_write: Optional[Callable[[redis.Redis, Dict[str, Any], str], None]] = None,
    on_error: Optional[Callable[[Exception], None]] = None,
    require_label: Optional[str] = None,
) -> None:
    if not data:
        return

    _require_configured(require_label or cache_key)

    fetched_at_iso = now_utc_iso()
    payload = {"data": data, "fetched_at": fetched_at_iso}

    try:
        redis_client = _config_redis()
    except Exception as exc:
        if on_error:
            on_error(exc)
        redis_client = None

    if redis_client is not None:
        try:
            redis_setex_json(redis_client, cache_key, ttl, payload)
        except Exception as exc:
            if on_error:
                on_error(exc)
        else:
            try:
                redis_set_json(redis_client, f"{cache_key}:last_success", payload)
            except Exception:
                pass
            if on_redis_write is not None:
                try:
                    on_redis_write(redis_client, data, fetched_at_iso)
                except Exception as exc:
                    if on_error:
                        on_error(exc)

    update_local_cache(local_cache, data, ttl, stale_grace, fetched_at_iso)


_TValue = TypeVar("_TValue")


def _refresh_with_backoff(
    *,
    load_cached: Callable[[bool], Tuple[Optional[_TValue], Dict[str, Any]]],
    fetcher: Callable[[], Optional[_TValue]],
    cache_writer: Callable[[_TValue], None],
    get_failure_until: Callable[[], float],
    set_failure_until: Callable[[float], None],
    allow_stale: bool = True,
    fallback_formatter: Optional[
        Callable[[Optional[_TValue], Dict[str, Any]], Optional[_TValue]]
    ] = None,
    on_cache_error: Optional[Callable[[Exception], None]] = None,
    backoff_window: int = TTL_BCRA,
) -> Optional[_TValue]:
    cached, meta = load_cached(allow_stale)
    if cached and meta.get("is_fresh"):
        return cached

    fallback = fallback_formatter(cached, meta) if fallback_formatter else cached

    now_ts = time.time()
    if now_ts < get_failure_until():
        return fallback

    fetched = fetcher()
    if fetched:
        try:
            cache_writer(fetched)
        except Exception as exc:
            if on_cache_error:
                on_cache_error(exc)
        set_failure_until(0.0)
        return fetched

    set_failure_until(now_ts + min(backoff_window, 120))
    return fallback


__all__ = [
    "CACHE_STALE_GRACE_BANDS",
    "CACHE_STALE_GRACE_BCRA",
    "TTL_BCRA",
    "TTL_MAYORISTA_MISSING",
    "bcra_api_get",
    "bcra_fetch_latest_variables",
    "bcra_get_value_for_date",
    "bcra_list_variables",
    "cache_bcra_variables",
    "cache_currency_band_limits",
    "cache_mayorista_missing",
    "get_country_risk_summary",
    "calculate_tcrm_100",
    "configure",
    "fetch_currency_band_limits",
    "format_bcra_variables",
    "get_cached_bcra_variables",
    "get_cached_tcrm_100",
    "get_currency_band_limits",
    "get_latest_itcrm_details",
    "get_latest_itcrm_value",
    "get_latest_itcrm_value_and_date",
    "get_or_refresh_bcra_variables",
    "reset_local_caches",
]


def bcra_api_get(
    path: str, params: Optional[Dict[str, Any]] = None, ttl: int = TTL_BCRA
) -> Optional[Dict[str, Any]]:
    """Call BCRA Estadísticas Monetarias v4.0 API."""

    _require_configured("bcra_api_get")

    try:
        base_url = "https://api.bcra.gob.ar/estadisticas/v4.0"
        url = base_url + (path if path.startswith("/") else "/" + path)
        warnings.filterwarnings("ignore", category=InsecureRequestWarning)
        resp = _call_cached_requests(
            url,
            params,
            None,
            ttl,
            hourly_cache=False,
            get_history=False,
            verify_ssl=False,
        )
        if not resp or "data" not in resp:
            return None
        data = resp["data"]
        if isinstance(data, dict):
            return data
        return None
    except Exception:
        return None


def bcra_list_variables(
    category: Optional[str] = "Principales Variables",
) -> Optional[List[Dict[str, Any]]]:
    """Return list of variables. If `category` is provided, filter client-side."""

    params: Optional[Dict[str, Any]] = None
    if category:
        params = {"limit": "2000"}
    data = bcra_api_get("/monetarias", params)
    try:
        if not data:
            return None
        results = data.get("results")
        if not isinstance(results, list):
            return None
        if not category:
            return results

        cat = _normalize_text(category)
        return [
            r for r in results if cat in _normalize_text(r.get("categoria", ""))
        ]
    except Exception:
        return None


def _parse_currency_band_rows(
    rows: Iterable[Iterable[Union[str, float, int, Decimal]]],
    *,
    today: Optional[date] = None,
) -> Optional[Dict[str, Any]]:
    """Return latest band row parsed into floats and formatted date."""

    effective_today = today or datetime.now(BA_TZ).date()
    parsed_rows: List[Tuple[date, float, float]] = []

    for row in rows:
        row_list = list(row)
        if len(row_list) < 3:
            continue
        date_raw, lower_raw, upper_raw = row_list[0], row_list[1], row_list[2]
        dt = parse_date_string(str(date_raw))
        if not dt:
            continue
        lower = parse_monetary_number(lower_raw)
        upper = parse_monetary_number(upper_raw)
        if lower is None or upper is None:
            continue
        parsed_rows.append((dt.date(), float(lower), float(upper)))

    if not parsed_rows:
        return None

    parsed_rows.sort(key=lambda item: item[0])

    on_or_before = [item for item in parsed_rows if item[0] <= effective_today]
    if on_or_before:
        current = on_or_before[-1]
        previous = on_or_before[-2] if len(on_or_before) > 1 else None
    else:
        current = parsed_rows[-1]
        previous = parsed_rows[-2] if len(parsed_rows) > 1 else None

    current_date, lower_val, upper_val = current

    def pct_change(
        curr: float, prev: Optional[Tuple[date, float, float]], index: int
    ) -> Optional[float]:
        if prev is None:
            return None
        prev_value_raw = prev[index]
        if not isinstance(prev_value_raw, (int, float)):
            return None
        prev_value = float(prev_value_raw)
        if prev_value == 0:
            return None
        try:
            return ((curr - prev_value) / prev_value) * 100
        except Exception:
            return None

    lower_change = pct_change(lower_val, previous, 1)
    upper_change = pct_change(upper_val, previous, 2)

    date_iso = current_date.isoformat()
    result: Dict[str, Any] = {
        "date": to_ddmmyy(date_iso),
        "date_iso": date_iso,
        "lower": lower_val,
        "upper": upper_val,
    }

    if lower_change is not None:
        result["lower_change_pct"] = lower_change
    if upper_change is not None:
        result["upper_change_pct"] = upper_change

    return result


def _fetch_currency_band_series(
    var_id: int,
    limit: int = 200,
    *,
    api_get_fn: Optional[
        Callable[[str, Optional[Dict[str, Any]]], Optional[Dict[str, Any]]]
    ] = None,
) -> Dict[date, float]:
    """Return a {date: value} map for the requested Principales Variables series."""

    api_get = api_get_fn or bcra_api_get

    data = api_get(f"/monetarias/{var_id}", {"limit": str(int(limit))})
    if not data:
        return {}
    results = data.get("results")
    if not isinstance(results, list):
        return {}

    series: Dict[date, float] = {}
    for entry in results:
        detalle = entry.get("detalle") if isinstance(entry, dict) else None
        if not isinstance(detalle, list):
            continue
        for row in detalle:
            if not isinstance(row, dict):
                continue
            date_raw = row.get("fecha")
            value_raw = row.get("valor")
            if value_raw is None:
                continue
            dt = parse_date_string(str(date_raw))
            if not dt:
                continue
            parsed_val = parse_monetary_number(value_raw)
            if parsed_val is None:
                continue
            series[dt.date()] = float(parsed_val)
    return series


def _get_cached_currency_band_entry(
    allow_stale: bool = False,
) -> Tuple[Optional[Dict[str, Any]], Dict[str, Any]]:
    """Return cached currency band limits with freshness metadata."""

    return _load_cached_entry(
        cache_key="bcra_currency_band_limits",
        local_cache=_currency_band_local_cache,
        ttl=TTL_BCRA,
        stale_grace=CACHE_STALE_GRACE_BANDS,
        allow_stale=allow_stale,
        redis_error_message="Error getting cached currency bands",
        require_label="_get_cached_currency_band_entry",
    )


def cache_currency_band_limits(data: Dict[str, Any], ttl: int = TTL_BCRA) -> None:
    if not data:
        return

    _persist_cache_entry(
        cache_key="bcra_currency_band_limits",
        data=data,
        local_cache=_currency_band_local_cache,
        ttl=ttl,
        stale_grace=CACHE_STALE_GRACE_BANDS,
        require_label="cache_currency_band_limits",
    )


def fetch_currency_band_limits(
    *,
    list_variables_fn: Optional[
        Callable[[Optional[str]], Optional[List[Dict[str, Any]]]]
    ] = None,
    api_get_fn: Optional[
        Callable[[str, Optional[Dict[str, Any]]], Optional[Dict[str, Any]]]
    ] = None,
) -> Optional[Dict[str, Any]]:
    """Fetch latest currency band limits (piso/techo)."""

    try:
        list_variables = list_variables_fn or bcra_list_variables
        api_get = api_get_fn or bcra_api_get

        variables = list_variables("Principales Variables")
        if not variables:
            return None

        lower_id: Optional[int] = None
        upper_id: Optional[int] = None

        for item in variables:
            if lower_id and upper_id:
                break
            if not isinstance(item, dict):
                continue
            desc = str(item.get("descripcion", ""))
            if not desc:
                continue
            normalized = _normalize_text(desc)
            if "bandas cambiarias" not in normalized:
                continue
            var_id = item.get("idVariable")
            if not isinstance(var_id, int):
                continue
            if "superior" in normalized and upper_id is None:
                upper_id = var_id
            elif "inferior" in normalized and lower_id is None:
                lower_id = var_id

        if lower_id is None or upper_id is None:
            return None

        lower_series = _fetch_currency_band_series(lower_id, api_get_fn=api_get)
        upper_series = _fetch_currency_band_series(upper_id, api_get_fn=api_get)
        if not lower_series or not upper_series:
            return None

        common_dates = sorted(set(lower_series.keys()) & set(upper_series.keys()))
        if not common_dates:
            return None

        max_rows = 180
        selected_dates = common_dates[-max_rows:]
        rows: List[List[Union[str, float]]] = [
            [
                dt.isoformat(),
                lower_series[dt],
                upper_series[dt],
            ]
            for dt in selected_dates
        ]

        return _parse_currency_band_rows(rows)
    except Exception as exc:
        print(f"Error fetching BCRA currency bands via API: {exc}")
        return None


def get_currency_band_limits(
    fetcher: Optional[Callable[[], Optional[Dict[str, Any]]]] = None,
) -> Optional[Dict[str, Any]]:
    """Return cached currency band limits, fetching and caching if necessary."""

    effective_fetcher = fetcher or fetch_currency_band_limits

    def load_cached(allow_stale: bool) -> Tuple[Optional[Dict[str, Any]], Dict[str, Any]]:
        return _get_cached_currency_band_entry(allow_stale=allow_stale)

    return _refresh_with_backoff(
        load_cached=load_cached,
        fetcher=effective_fetcher,
        cache_writer=lambda data: cache_currency_band_limits(data, TTL_BCRA),
        get_failure_until=_get_currency_band_failure_until,
        set_failure_until=_set_currency_band_failure_until,
    )


def bcra_fetch_latest_variables() -> Optional[Dict[str, Dict[str, str]]]:
    """Fetch latest Principales Variables via BCRA API."""

    try:
        vars_list = bcra_list_variables("Principales Variables")
        if not vars_list:
            return None
        variables: Dict[str, Dict[str, str]] = {}
        for item in vars_list:
            name = str(item.get("descripcion", "")).strip()
            date_iso = str(item.get("ultFechaInformada", "")).strip()
            val = item.get("ultValorInformado", None)
            if not name or val is None or not date_iso:
                continue
            variables[name] = {"value": to_es_number(val), "date": to_ddmmyy(date_iso)}
        return variables
    except Exception as exc:
        print(f"Error fetching BCRA variables via API: {exc}")
        return None


def bcra_get_value_for_date(desc_substr: str, date_iso: str) -> Optional[float]:
    """Get numeric value for the variable matching `desc_substr` on `date_iso`."""

    try:
        vars_list = bcra_list_variables("Principales Variables")
        if not vars_list:
            return None

        target = _normalize_text(desc_substr)
        var_id: Optional[int] = None
        var_name: Optional[str] = None
        for entry in vars_list:
            desc = str(entry.get("descripcion", ""))
            if target in _normalize_text(desc):
                vid = entry.get("idVariable")
                if isinstance(vid, int):
                    var_id = vid
                    var_name = desc
                    break
        if var_id is None:
            return None

        data = bcra_api_get(
            f"/monetarias/{var_id}",
            {"desde": date_iso, "hasta": date_iso, "limit": "1"},
        )
        if not data:
            return None
        results = data.get("results")
        if not isinstance(results, list) or not results:
            return None
        detalle = results[0].get("detalle")
        if not isinstance(detalle, list) or not detalle:
            return None
        row = detalle[0]
        valor = row.get("valor")
        if valor is None:
            return None
        try:
            val_f = float(valor)
        except Exception:
            return None

        try:
            if var_name and re.search(
                "tipo.*cambio.*mayorista", _normalize_text(var_name)
            ):
                redis_client = _config_redis()
                redis_set_json(
                    redis_client,
                    f"bcra_mayorista:{date_iso}",
                    {"value": val_f, "date": to_ddmmyy(date_iso)},
                )
        except Exception:
            pass
        return val_f
    except Exception:
        return None


def _get_cached_bcra_cache_entry(
    allow_stale: bool = False,
) -> Tuple[Optional[Dict[str, Any]], Dict[str, Any]]:
    """Return cached BCRA variables along with freshness metadata."""

    return _load_cached_entry(
        cache_key="bcra_variables",
        local_cache=_bcra_local_cache,
        ttl=TTL_BCRA,
        stale_grace=CACHE_STALE_GRACE_BCRA,
        allow_stale=allow_stale,
        require_label="_get_cached_bcra_cache_entry",
    )


def get_cached_bcra_variables(allow_stale: bool = False) -> Optional[Dict[str, Any]]:
    value, _ = _get_cached_bcra_cache_entry(allow_stale=allow_stale)
    return value


def cache_bcra_variables(variables: Dict[str, Any], ttl: int = TTL_BCRA) -> None:
    if not variables:
        return

    def _log_cache_error(exc: Exception) -> None:
        print(f"Error caching BCRA variables: {exc}")

    def _store_mayorista(
        redis_client: redis.Redis, payload_data: Dict[str, Any], _: str
    ) -> None:
        for key, raw_data in (payload_data or {}).items():
            normalized_key = _normalize_text(key)
            if not re.search("tipo.*cambio.*mayorista", normalized_key):
                continue
            if not isinstance(raw_data, Mapping):
                continue
            value_num = parse_monetary_number(raw_data.get("value", ""))
            parsed_dt = parse_date_string(str(raw_data.get("date", "")))
            if value_num is None or not parsed_dt:
                continue
            date_key = parsed_dt.date().isoformat()
            redis_set_json(
                redis_client,
                f"bcra_mayorista:{date_key}",
                {"value": value_num, "date": str(raw_data.get("date", ""))},
            )
            break

    _persist_cache_entry(
        cache_key="bcra_variables",
        data=variables,
        local_cache=_bcra_local_cache,
        ttl=ttl,
        stale_grace=CACHE_STALE_GRACE_BCRA,
        on_redis_write=_store_mayorista,
        on_error=_log_cache_error,
        require_label="cache_bcra_variables",
    )


def _attach_bcra_meta(value: Optional[Dict[str, Any]], meta: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if not value:
        return None
    if meta.get("is_fresh"):
        return value

    enriched = dict(value)
    meta_payload: Dict[str, Any] = {"stale": True}
    fetched_at = meta.get("fetched_at")
    if fetched_at:
        meta_payload["fetched_at"] = fetched_at
    enriched["_meta"] = meta_payload
    return enriched


def get_or_refresh_bcra_variables() -> Optional[Dict[str, Any]]:
    """Return BCRA variables using cache or API, preserving stale fallback."""

    return _refresh_with_backoff(
        load_cached=lambda allow_stale: _get_cached_bcra_cache_entry(
            allow_stale=allow_stale
        ),
        fetcher=bcra_fetch_latest_variables,
        cache_writer=cache_bcra_variables,
        get_failure_until=_get_bcra_failure_until,
        set_failure_until=_set_bcra_failure_until,
        fallback_formatter=_attach_bcra_meta,
    )


def get_latest_itcrm_value() -> Optional[float]:
    try:
        details = get_latest_itcrm_value_and_date()
        return details[0] if details else None
    except Exception:
        return None


def _get_itcrm_value_for_date(target_dt: datetime) -> Optional[Tuple[float, datetime]]:
    """Return ITCRM value and sheet date for the closest entry on/before `target_dt`."""

    try:
        warnings.filterwarnings("ignore", category=InsecureRequestWarning)
        response = requests.get(
            "https://www.bcra.gob.ar/Pdfs/PublicacionesEstadisticas/ITCRMSerie.xlsx",
            timeout=10,
            verify=False,
        )
        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        sheet_like = getattr(workbook, "active", None) or (
            workbook.worksheets[0] if getattr(workbook, "worksheets", None) else None
        )
        if sheet_like is None:
            return None
        sheet = cast(Any, sheet_like)

        def normalize_date(value: Any) -> Optional[datetime]:
            try:
                if isinstance(value, datetime):
                    return value
                if isinstance(value, date):
                    return datetime.combine(value, datetime.min.time())
                if hasattr(value, "date"):
                    maybe_date = value.date()
                    if isinstance(maybe_date, date):
                        return datetime.combine(maybe_date, datetime.min.time())
            except Exception:
                pass

            parsed = parse_date_string(str(value or ""))
            if parsed:
                return parsed
            return None

        def normalize_value(cell_value: Any) -> Optional[float]:
            if cell_value is None:
                return None
            if isinstance(cell_value, (int, float, Decimal)):
                return float(cell_value)
            try:
                text_val = str(cell_value).strip()
            except Exception:
                return None
            if not text_val:
                return None
            try:
                return float(text_val)
            except ValueError:
                try:
                    normalized = text_val.replace(".", "").replace(",", ".")
                    return float(normalized)
                except Exception:
                    return None

        target_date = target_dt.date()
        for row in range(sheet.max_row, 0, -1):
            row_date_raw = sheet.cell(row=row, column=1).value
            row_value_raw = sheet.cell(row=row, column=2).value
            row_date_dt = normalize_date(row_date_raw)
            if row_date_dt is None:
                continue
            row_date_only = row_date_dt.date()
            if row_date_only > target_date:
                continue
            value = normalize_value(row_value_raw)
            if value is None:
                continue
            normalized_dt = datetime.combine(row_date_only, datetime.min.time())
            return value, normalized_dt

        return None
    except Exception as exc:
        print(f"Error finding ITCRM value for {target_dt.date()}: {exc}")
        return None


def cache_mayorista_missing(
    date_key: str,
    redis_client: Optional[redis.Redis] = None,
    *,
    redis_setex_json_fn: Optional[
        Callable[[redis.Redis, str, int, Any], Any]
    ] = None,
) -> None:
    """Store a short-lived sentinel indicating mayorista is missing for date_key."""

    try:
        client = redis_client or _config_redis()
        if not client:
            return
        payload = {"missing": True, "timestamp": int(time.time())}
        setex_json = redis_setex_json_fn or redis_setex_json
        setex_json(
            client,
            f"bcra_mayorista:{date_key}",
            TTL_MAYORISTA_MISSING,
            payload,
        )
    except Exception:
        pass


def calculate_tcrm_100(
    target_date: Optional[Union[str, datetime, date]] = None,
    *,
    config_redis_fn: Optional[Callable[..., redis.Redis]] = None,
    redis_get_json_fn: Optional[Callable[[redis.Redis, str], Any]] = None,
    redis_set_json_fn: Optional[Callable[[redis.Redis, str, Any], Any]] = None,
    redis_setex_json_fn: Optional[
        Callable[[redis.Redis, str, int, Any], Any]
    ] = None,
    bcra_get_value_for_date_fn: Optional[Callable[[str, str], Optional[float]]] = None,
    cache_mayorista_missing_fn: Optional[
        Callable[[str, Optional[redis.Redis]], None]
    ] = None,
    itcrm_getter_fn: Optional[Callable[[], Optional[Tuple[float, str]]]] = None,
) -> Optional[float]:
    """Calculate nominal exchange rate that sets ITCRM to 100."""

    try:
        config = config_redis_fn or _config_redis
        get_json = redis_get_json_fn or redis_get_json
        set_json = redis_set_json_fn or redis_set_json
        bcra_value_for_date = bcra_get_value_for_date_fn or bcra_get_value_for_date
        cache_missing = cache_mayorista_missing_fn or cache_mayorista_missing
        itcrm_getter = itcrm_getter_fn or get_latest_itcrm_value_and_date

        normalized_target: Optional[datetime] = None
        if target_date is not None:
            if isinstance(target_date, datetime):
                normalized_target = target_date
            elif isinstance(target_date, date):
                normalized_target = datetime.combine(target_date, datetime.min.time())
            else:
                normalized_target = parse_date_string(str(target_date))
            if normalized_target is None:
                return None

        itcrm_value: Optional[float]
        itcrm_dt: Optional[datetime]

        if normalized_target is None:
            details = itcrm_getter()
            if not details:
                return None
            raw_value, raw_date_str = details
            try:
                itcrm_value = float(raw_value)
            except Exception:
                return None
            parsed_dt = parse_date_string(raw_date_str)
            if not parsed_dt:
                return None
            itcrm_dt = parsed_dt
        else:
            lookup = _get_itcrm_value_for_date(normalized_target)
            if not lookup:
                return None
            value, matched_dt = lookup
            try:
                itcrm_value = float(value)
            except Exception:
                return None
            itcrm_dt = matched_dt

        if itcrm_dt is None:
            return None

        date_key = itcrm_dt.date().isoformat()

        try:
            redis_client = config()
        except Exception:
            redis_client = None

        wholesale_value: Optional[float] = None
        if redis_client is not None:
            try:
                cached = get_json(redis_client, f"bcra_mayorista:{date_key}")
                if isinstance(cached, dict):
                    if cached.get("missing"):
                        return None
                    if "value" in cached:
                        parsed = parse_monetary_number(cached["value"])
                        if parsed is not None:
                            wholesale_value = float(parsed)
                if wholesale_value is None:
                    fetched = bcra_value_for_date(
                        "tipo de cambio mayorista", date_key
                    )
                    if fetched is not None:
                        wholesale_value = float(fetched)
                    else:
                        cache_missing(
                            date_key,
                            redis_client,
                            redis_setex_json_fn=redis_setex_json_fn,
                        )
                        return None
            except Exception:
                pass

        if wholesale_value is None:
            fetched = bcra_value_for_date("tipo de cambio mayorista", date_key)
            if fetched is not None:
                wholesale_value = float(fetched)
            else:
                cache_missing(
                    date_key,
                    redis_client,
                    redis_setex_json_fn=redis_setex_json_fn,
                )
                return None

        result = wholesale_value * 100 / itcrm_value

        if redis_client is not None and wholesale_value is not None:
            try:
                set_json(
                    redis_client,
                    f"bcra_mayorista:{date_key}",
                    {"value": wholesale_value, "date": to_ddmmyy(date_key)},
                )
            except Exception:
                pass
        return result
    except Exception as exc:
        print(f"Error calculating TCRM 100: {exc}")
        return None


def get_cached_tcrm_100(
    hours_ago: int = 24,
    expiration_time: int = 300,
    *,
    config_redis_fn: Optional[Callable[..., redis.Redis]] = None,
    redis_get_json_fn: Optional[Callable[[redis.Redis, str], Any]] = None,
    redis_set_json_fn: Optional[Callable[[redis.Redis, str, Any], Any]] = None,
    redis_setex_json_fn: Optional[
        Callable[[redis.Redis, str, int, Any], Any]
    ] = None,
    calculate_tcrm_fn: Optional[Callable[..., Optional[float]]] = None,
    get_latest_itcrm_fn: Optional[Callable[[], Optional[Tuple[float, str]]]] = None,
    bcra_get_value_for_date_fn: Optional[Callable[[str, str], Optional[float]]] = None,
    cache_mayorista_missing_fn: Optional[
        Callable[[str, Optional[redis.Redis]], None]
    ] = None,
    get_cache_history_fn: Optional[
        Callable[[int, str, redis.Redis], Optional[Dict[str, Any]]]
    ] = None,
) -> Tuple[Optional[float], Optional[float]]:
    """Get cached TCRM 100 value with optional historical change."""

    cache_key = "tcrm_100"

    try:
        config = config_redis_fn or _config_redis
        get_json = redis_get_json_fn or redis_get_json
        set_json = redis_set_json_fn or redis_set_json
        setex_json = redis_setex_json_fn
        calculate_tcrm = calculate_tcrm_fn or calculate_tcrm_100
        latest_itcrm = get_latest_itcrm_fn or get_latest_itcrm_value_and_date
        bcra_value_for_date = bcra_get_value_for_date_fn or bcra_get_value_for_date
        cache_missing = cache_mayorista_missing_fn or cache_mayorista_missing
        get_history = get_cache_history_fn or _get_cache_history

        redis_client = config()
        redis_response = get_json(redis_client, cache_key)
        timestamp = int(time.time())

        same_day_ok = False
        skip_mayorista_fetch = False
        try:
            itcrm_cached = redis_get_json(redis_client, "latest_itcrm_details")
            itcrm_date_str = None
            if isinstance(itcrm_cached, dict) and "date" in itcrm_cached:
                itcrm_date_str = str(itcrm_cached.get("date", ""))
            else:
                details = latest_itcrm()
                if details:
                    itcrm_date_str = details[1]
            dt = parse_date_string(itcrm_date_str or "")
            if dt is not None:
                date_key = dt.date().isoformat()
                mayorista_cached = get_json(
                    redis_client, f"bcra_mayorista:{date_key}"
                )
                if isinstance(mayorista_cached, dict):
                    if mayorista_cached.get("missing"):
                        skip_mayorista_fetch = True
                    elif "value" in mayorista_cached:
                        if parse_monetary_number(mayorista_cached["value"]) is not None:
                            same_day_ok = True
                if not same_day_ok and not skip_mayorista_fetch:
                    fetched_val = bcra_value_for_date(
                        "tipo de cambio mayorista", date_key
                    )
                    if fetched_val is not None:
                        same_day_ok = True
                    else:
                        cache_missing(
                            date_key,
                            redis_client,
                            redis_setex_json_fn=setex_json,
                        )
                        skip_mayorista_fetch = True
        except Exception:
            same_day_ok = False

        history_data: Optional[Dict[str, Any]] = None
        try:
            history_data = get_history(hours_ago, cache_key, redis_client)
        except Exception:
            history_data = None

        if history_data is None and hours_ago and not skip_mayorista_fetch:
            try:
                history_dt = datetime.now(BA_TZ) - timedelta(hours=hours_ago)
                history_prefix = history_dt.strftime("%Y-%m-%d-%H")
                history_key = history_prefix + cache_key
                existing_slot = redis_client.get(history_key)
                if not existing_slot:
                    backfill_value = calculate_tcrm(target_date=history_dt)
                    if backfill_value is not None:
                        history_timestamp = int(
                            history_dt.replace(
                                minute=0, second=0, microsecond=0
                            ).timestamp()
                        )
                        history_data = {
                            "timestamp": history_timestamp,
                            "data": backfill_value,
                        }
                        set_json(redis_client, history_key, history_data)
            except Exception:
                history_data = None

        history_value = history_data["data"] if history_data else None

        def compute_and_store() -> Optional[float]:
            value = calculate_tcrm()
            if value is None:
                return None
            redis_value = {"timestamp": timestamp, "data": value}
            set_json(redis_client, cache_key, redis_value)
            current_hour = datetime.now(BA_TZ).strftime("%Y-%m-%d-%H")
            set_json(redis_client, current_hour + cache_key, redis_value)
            return value

        if not same_day_ok or skip_mayorista_fetch:
            current_value = None
        elif redis_response is None:
            current_value = compute_and_store()
        else:
            cached_data = cast(Dict[str, Any], redis_response)
            cache_age = timestamp - int(cached_data["timestamp"])
            if cache_age > expiration_time:
                current_value = compute_and_store()
            else:
                current_value = cached_data.get("data")

        change = None
        if current_value is not None and history_value:
            try:
                change = ((current_value / history_value) - 1) * 100
            except ZeroDivisionError:
                change = None

        return current_value, change
    except Exception as exc:
        print(f"Error getting cached TCRM 100: {exc}")
        return None, None


def get_latest_itcrm_details() -> Optional[Tuple[float, str]]:
    """Return latest TCRM value and its date (DD/MM/YY) from the spreadsheet."""

    try:
        warnings.filterwarnings("ignore", category=InsecureRequestWarning)
        response = requests.get(
            "https://www.bcra.gob.ar/Pdfs/PublicacionesEstadisticas/ITCRMSerie.xlsx",
            timeout=10,
            verify=False,
        )
        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        sheet_like = getattr(workbook, "active", None) or (
            workbook.worksheets[0] if getattr(workbook, "worksheets", None) else None
        )
        if sheet_like is None:
            return None
        sheet = cast(Any, sheet_like)

        def parse_date_cell(value: Any) -> Optional[str]:
            try:
                if hasattr(value, "strftime"):
                    try:
                        d = value.date() if hasattr(value, "date") else value
                        return f"{d.day:02d}/{d.month:02d}/{d.year % 100:02d}"
                    except Exception:
                        pass
            except Exception:
                pass
            try:
                string_val = str(value).strip()
                for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
                    try:
                        dt = datetime.strptime(string_val, fmt)
                        return f"{dt.day:02d}/{dt.month:02d}/{dt.year % 100:02d}"
                    except Exception:
                        continue
            except Exception:
                pass
            return None

        for row in range(sheet.max_row, 0, -1):
            date_cell = sheet.cell(row=row, column=1).value
            cell_value = sheet.cell(row=row, column=2).value
            if cell_value is None:
                continue
            val: Optional[float] = None
            if isinstance(cell_value, (int, float, Decimal)):
                val = float(cell_value)
            else:
                text_val = str(cell_value)
                try:
                    val = float(text_val)
                except ValueError:
                    try:
                        normalized = text_val.replace(".", "").replace(",", ".")
                        val = float(normalized)
                    except Exception:
                        continue
            date_str = parse_date_cell(date_cell) or ""
            if val is not None:
                return val, date_str
        return None
    except Exception as exc:
        print(f"Error fetching ITCRM details: {exc}")
        return None


def get_latest_itcrm_value_and_date() -> Optional[Tuple[float, str]]:
    """Cached wrapper returning latest ITCRM value and its date."""

    try:
        redis_client = _config_redis()
        cached = redis_get_json(redis_client, "latest_itcrm_details")
        if isinstance(cached, dict) and "value" in cached:
            try:
                return float(cached["value"]), str(cached.get("date", ""))
            except Exception:
                pass
    except Exception:
        pass

    details = get_latest_itcrm_details()
    if not details:
        return None
    value, date_str = details
    try:
        redis_client = _config_redis()
        redis_setex_json(
            redis_client,
            "latest_itcrm_details",
            1800,
            {"value": value, "date": date_str},
        )
    except Exception:
        pass
    return value, date_str


def _parse_iso_datetime(value: Any) -> Optional[datetime]:
    if not isinstance(value, str) or not value:
        return None
    try:
        normalized = value.replace("Z", "+00:00")
        dt = datetime.fromisoformat(normalized)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(BA_TZ)
    except Exception:
        return None


def _fetch_country_risk_direct() -> Optional[Mapping[str, Any]]:
    """Fetch country risk directly from BondTerminal bypassing Redis cache."""

    try:
        warnings.filterwarnings("ignore", category=InsecureRequestWarning)
        resp = requests.get(
            "https://bondterminal.com/api/riesgo-pais",
            timeout=5,
            verify=False,
        )
        resp.raise_for_status()
        data = resp.json()
        if isinstance(data, Mapping):
            return data
    except Exception as exc:
        print(f"Fallback request for country risk failed: {exc}")
    return None


def get_country_risk_summary() -> Optional[Dict[str, Any]]:
    """Fetch and normalize Argentine country risk from BondTerminal."""

    response_payload: Optional[Mapping[str, Any]] = None

    try:
        response = _call_cached_requests(
            "https://bondterminal.com/api/riesgo-pais",
            None,
            None,
            TTL_COUNTRY_RISK,
            verify_ssl=False,
        )
        if isinstance(response, Mapping):
            candidate_payload = response.get("data")
            if isinstance(candidate_payload, Mapping):
                response_payload = candidate_payload
            else:
                print("Country risk payload missing or invalid")
        else:
            print(
                "Unexpected response while fetching country risk: "
                f"{type(response).__name__}"
            )
    except Exception as exc:
        print(f"Error requesting country risk data: {exc}")

    if response_payload is None:
        response_payload = _fetch_country_risk_direct()
        if response_payload is None:
            return None

    value = response_payload.get("weightedSpreadBps")
    try:
        value_bps = float(value)
    except Exception as exc:
        print(f"Invalid country risk value: {value} ({exc})")
        return None

    delta_raw = None
    deltas = response_payload.get("deltas")
    if isinstance(deltas, Mapping):
        delta_raw = deltas.get("oneDay")

    delta_value: Optional[float] = None
    if delta_raw is not None:
        try:
            delta_value = float(delta_raw)
        except Exception as exc:
            print(f"Invalid delta for country risk: {delta_raw} ({exc})")
            delta_value = None

    valuation_dt: Optional[datetime] = None
    for key in ("valuationDate", "asOf", "lastDataTickIso"):
        try:
            valuation_dt = _parse_iso_datetime(response_payload.get(key))
        except Exception as exc:
            print(
                "Error parsing country risk valuation datetime "
                f"from key '{key}': {exc}"
            )
            valuation_dt = None
        if valuation_dt:
            break

    label: Optional[str] = None
    if valuation_dt:
        label = valuation_dt.strftime("%d/%m %H:%M")

    return {
        "value_bps": value_bps,
        "delta_one_day": delta_value,
        "valuation_datetime": valuation_dt,
        "valuation_label": label,
    }


def format_bcra_variables(
    variables: Dict[str, Any],
    *,
    band_fetcher: Optional[Callable[[], Optional[Dict[str, Any]]]] = None,
    itcrm_getter: Optional[Callable[[], Optional[Tuple[float, str]]]] = None,
    country_risk_getter: Optional[Callable[[], Optional[Dict[str, Any]]]] = None,
) -> str:
    """Format BCRA variables for display (robust to naming changes)."""

    if not variables:
        return "No se pudieron obtener las variables del BCRA"

    get_bands = band_fetcher or get_currency_band_limits
    get_itcrm = itcrm_getter or get_latest_itcrm_value_and_date
    get_country_risk = country_risk_getter or get_country_risk_summary

    def format_value(value_str: str, is_percentage: bool = False) -> str:
        try:
            clean_value = (
                value_str.replace(".", "").replace(",", ".")
                if not is_percentage
                else value_str.replace(",", ".")
            )
            num = float(clean_value)
            if is_percentage:
                return f"{num:.1f}%" if num >= 10 else f"{num:.2f}%"
            if num >= 1_000_000:
                return f"{num/1000:,.0f}".replace(",", ".")
            if num >= 1000:
                return f"{num:,.0f}".replace(",", ".")
            return f"{num:.2f}".replace(".", ",")
        except Exception:
            return f"{value_str}%" if is_percentage else value_str

    specs = [
        (
            r"base\s*monetaria",
            lambda v: f"🏦 Base monetaria: ${format_value(v)} mill. pesos",
        ),
        (
            r"variacion.*mensual.*indice.*precios.*consumidor|inflacion.*mensual",
            lambda v: f"📈 Inflación mensual: {format_value(v, True)}",
        ),
        (
            r"variacion.*interanual.*indice.*precios.*consumidor|inflacion.*interanual",
            lambda v: f"📊 Inflación interanual: {format_value(v, True)}",
        ),
        (
            r"(mediana.*variacion.*interanual.*(12|doce).*meses.*(relevamiento.*expectativas.*mercado|rem)|inflacion.*esperada)",
            lambda v: f"🔮 Inflación esperada: {format_value(v, True)}",
        ),
        (r"tamar", lambda v: f"📈 TAMAR: {format_value(v, True)}"),
        (r"badlar", lambda v: f"📊 BADLAR: {format_value(v, True)}"),
        (
            r"tipo.*cambio.*minorista|minorista.*promedio.*vendedor",
            lambda v: f"💵 Dólar minorista: ${v}",
        ),
        (r"tipo.*cambio.*mayorista", lambda v: f"💱 Dólar mayorista: ${v}"),
        (r"unidad.*valor.*adquisitivo|\buva\b", lambda v: f"💰 UVA: ${v}"),
        (r"coeficiente.*estabilizacion.*referencia|\bcer\b", lambda v: f"📊 CER: {v}"),
        (
            r"reservas.*internacionales",
            lambda v: f"🏛️ Reservas: USD {format_value(v)} millones",
        ),
    ]

    meta_info: Dict[str, Any] = {}
    if isinstance(variables, dict):
        candidate_meta = variables.get("_meta")
        if isinstance(candidate_meta, dict):
            meta_info = candidate_meta

    lines = ["📊 Variables principales BCRA\n"]
    latest_dt: Optional[datetime] = None
    for pattern, formatter in specs:
        compiled = re.compile(pattern)
        for key, data in variables.items():
            if str(key).startswith("_"):
                continue
            if not isinstance(data, Mapping):
                continue
            if compiled.search(_normalize_text(key)):
                value = data.get("value", "")
                date_label = data.get("date", "")
                line = formatter(value)
                if date_label and date_label != value:
                    line += f" ({str(date_label).replace('/2025', '/25')})"
                lines.append(line)
                parsed_dt = parse_date_string(str(date_label))
                if parsed_dt and (latest_dt is None or parsed_dt > latest_dt):
                    latest_dt = parsed_dt
                break

    try:
        country_risk = get_country_risk()
    except Exception:
        country_risk = None

    if country_risk:
        value_bps = country_risk.get("value_bps")
        if isinstance(value_bps, (int, float)):
            value_decimals = 1 if abs(value_bps) < 100 else 0
            value_text = fmt_num(float(value_bps), value_decimals).replace(".", ",")
            risk_line = f"🇦🇷 Riesgo país: {value_text} bps"

            details: List[str] = []
            label = country_risk.get("valuation_label")
            if isinstance(label, str) and label:
                details.append(label)

            delta_value = country_risk.get("delta_one_day")
            if isinstance(delta_value, (int, float)):
                abs_delta = abs(delta_value)
                if abs_delta >= 0.05:
                    delta_decimals = 1 if abs_delta < 100 else 0
                    delta_text = fmt_num(abs_delta, delta_decimals).replace(".", ",")
                    sign = "+" if delta_value > 0 else "-"
                    details.append(f"{sign}{delta_text} bps vs ayer")

            if details:
                risk_line += " (" + " | ".join(details) + ")"

            lines.append(risk_line)

    band_limits = get_bands()
    if band_limits:
        lower = band_limits.get("lower")
        upper = band_limits.get("upper")
        if isinstance(lower, (int, float)) and isinstance(upper, (int, float)):
            date_label = band_limits.get("date")
            lower_text = fmt_num(float(lower), 2)
            upper_text = fmt_num(float(upper), 2)
            line = f"📏 Bandas cambiarias: piso ${lower_text} / techo ${upper_text}"
            if isinstance(date_label, str) and date_label:
                line += f" ({date_label})"
            lines.append(line)

    try:
        details = get_itcrm()
        if details:
            itcrm_value, date_str = details
            lines.append(
                f"📐 TCRM: {fmt_num(float(itcrm_value), 2)}"
                + (f" ({date_str})" if date_str else "")
            )
    except Exception:
        pass

    if meta_info.get("stale"):
        stale_msg = (
            "⚠️ No hay actualización nueva del BCRA, te muestro lo último que tengo."
        )
        if stale_msg not in lines:
            lines.append(stale_msg)

    if latest_dt:
        age_days = (datetime.now(BA_TZ).date() - latest_dt.date()).days
        if age_days >= 3:
            lines.append(
                f"⚠️ Datos del BCRA con {age_days} días de atraso, chequeá más tarde."
            )

    return "\n".join(lines)
